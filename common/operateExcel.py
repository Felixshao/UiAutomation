import os
import time
import pandas as pd
import xlrd
import openpyxl
from config.getProjectPath import get_project_path
from common.log import Logger

path = get_project_path()
log = Logger().get_logger()


class operateExcel():
    """
    操作excel类
    """
    def __init__(self, file_name=None, sheet_name=None):
        """
        :param file_name: 文件名称
        param file_name: 表名称
        """
        self.file_name = os.path.join(path, 'file', file_name)  # 连接文件路径
        self.sheet_name = sheet_name

    def get_excel_list(self):
        """
        读取excel并存入list
        :return:data;数据存放在list中返回
        """
        table = xlrd.open_workbook(self.file_name)
        sheet = table.sheet_by_name(self.sheet_name)
        data = [[sheet.cell_value(i, j).replace('\n', '') for j in range(sheet.ncols)]
                for i in range(1, sheet.nrows)]
        return data

    def get_excel_dict(self):
        """
        读取excel并存入dict
        :return:data;数据存放在dict中返回
        """
        if os.path.exists(self.file_name):
            table = xlrd.open_workbook(self.file_name)
        else:
            self.create_excel()
            table = xlrd.open_workbook(self.file_name)
        if self.sheet_name in table.sheet_names():
            sheet = table.sheet_by_name(self.sheet_name)
        else:
            self.create_sheet()
            table = xlrd.open_workbook(self.file_name)
            sheet = table.sheet_by_name(self.sheet_name)
        data = {}
        for i in range(2, sheet.nrows):
            data[sheet.cell_value(i, 0)] = sheet.cell_value(i, 1)
        return data

    def input_excel(self, result, row=1, col=1):
        """
        在excel写入结果
        :param result: 传入结果dict
        """
        try:
            table = openpyxl.load_workbook(self.file_name)
        except BaseException as per:
            log.error('当前表格已当前，请手动关闭后再试，文件:'.format(self.file_name))
            log.error(per)
            print('当前表格已当前，请手动关闭后再试，文件:', self.file_name)
            raise
        if self.sheet_name in table.sheetnames:
            sheet = table[self.sheet_name]  # 选择指定表
        else:
            self.create_sheet()
            table = openpyxl.load_workbook(self.file_name)
            sheet = table[self.sheet_name]  # 选择指定表
        row = 3
        for i in result:
            sheet.cell(row, 1, i)
            sheet.cell(row, 2, str(result[i]))
            row += 1
        table.save(self.file_name)

    def create_sheet(self):
        """表格中新增sheet表"""
        if os.access(self.file_name, os.R_OK):
            table = openpyxl.load_workbook(self.file_name)
            table.create_sheet(title=self.sheet_name)
            table.save(self.file_name)
        else:
            print('文件已打开，不可执行，请手动关闭表格后再试!')

    def create_excel(self):
        """新增xlsx表格"""
        table = openpyxl.Workbook()
        table.save(self.file_name)

    def get_exlce(self):
        """使用pandas库读取excel并存入dict"""
        df = pd.read_excel(self.file_name)
        res = df.to_dict(orient="records")
        print(res)
        return res


if __name__ == '__main__':

    read = operateExcel('stock1.xlsx', 'Sheet1')
    result = {'a': 1, 'b': 2}
    result2 = {'c': 3, 'd': 4}
    re = {}
    li = [1]
    data = read.get_exlce()
    print(data)
    # if li:
    #     print(li)
    # else:
    #     print(1)
    # read.input_excel(result)