import unittest, ddt, requests, urllib3, re, datetime, time, warnings, os
import openpyxl
from openpyxl.styles import colors, Font
from common.operateExcel import operateExcel
from common.log import Logger
from config.getProjectPath import get_project_path


warnings.filterwarnings('ignore')
honor_list = operateExcel(file_name='zc_interface.xlsx', sheet_name='list').get_excel_list()
log = Logger('testinterface.ZClistinter_test').get_logger()
path = get_project_path()


@ddt.ddt
class ZClistinter_test(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        log.info('------------------------  ZClistinter_test  开始 -------------------------------------------------')
        urllib3.disable_warnings()
        login = operateExcel(file_name='zc_interface.xlsx', sheet_name='login').get_excel_list()[0]
        cls.s = requests.session()
        cls.s.post(url=login[1], data=eval(login[3]), verify=False)
        cls.rank = list(datetime.datetime.now().isocalendar())      # 获取当前年份、第几周、第几天
        # 通过周数计算期数
        if (datetime.datetime.strptime(str(cls.rank[0]) + '-01-01', '%Y-%m-%d')).isocalendar()[2] != 1:
            cls.rank[1] = cls.rank[1] - 1

    @classmethod
    def tearDownClass(cls):
        log.info('------------------------ ZClistinter_test 结束 -------------------------------------------------')

    @ddt.data(honor_list[0])
    def test1_newhonor_list(self, data):
        """最新一期榜单"""
        data[3] = eval(data[3])
        data[4] = eval(data[4])
        if self.rank[1] - 1 <= 0:
            data[3]['rank'] = str(self.rank[0] - 1) + '52'
        else:
            data[3]['rank'] = str(self.rank[0]) + str(self.rank[1] - 1)
        result = self.check_hororlist_inter(data)
        self.input_excel_result(0 + 2, result)

    @ddt.data(honor_list[1])
    def test2_lastterm_list(self, data):
        """上一期榜单"""
        data[3] = eval(data[3])
        data[4] = eval(data[4])
        if self.rank[1] - 2 <= 0:
            data[3]['rank'] = str(self.rank[0] - 1) + str(52 + self.rank[1] - 2)
            data[0] = data[3]['rank'] + '期榜单'
        else:
            data[3]['rank'] = str(self.rank[0]) + str(self.rank[1] - 2)
            data[0] = data[3]['rank'] + '期榜单'
        result = self.check_hororlist_inter(data)
        self.input_excel_result(1 + 2, result)

    @ddt.data(honor_list[2])
    def test3_lastissue_list(self, data):
        """上上期榜单"""
        data[3] = eval(data[3])
        data[4] = eval(data[4])
        if self.rank[1] - 3 == 0 or self.rank[1] - 3 < 0:
            data[3]['rank'] = str(self.rank[0] - 1) + str(52 + self.rank[1] - 3)
            data[0] = data[3]['rank'] + '期榜单'
        else:
            data[3]['rank'] = str(self.rank[0]) + str(self.rank[1] - 3)
            data[0] = data[3]['rank'] + '期榜单'
        result = self.check_hororlist_inter(data)
        self.input_excel_result(2 + 2, result)

    @ddt.data(honor_list[3])
    def test4_lastperiod_list(self, data):
        """上上上期榜单"""
        title = data[0]
        data[3] = eval(data[3])
        data[4] = eval(data[4])
        if self.rank[1] - 4 == 0 or self.rank[1] - 4 < 0:
            data[3]['rank'] = str(self.rank[0] - 1) + str(52 + self.rank[1] - 4)
            data[0] = data[3]['rank'] + '期榜单'
        else:
            data[3]['rank'] = str(self.rank[0]) + str(self.rank[1] - 4)
            data[0] = data[3]['rank'] + '期榜单'
        result = self.check_hororlist_inter(data)
        self.input_excel_result(3+2, result)

    def check_hororlist_inter(self, data):
        """断言榜单接口"""
        try:
            honor = self.s.get(url=data[1], params=data[3], verify=False)
            self.assertEqual(honor.status_code, data[4]['status_code'],
                             msg='status_code:{0}'.format(honor.status_code))
            pattern = re.compile(pattern='<div class="dis_inlblo"><span class="rank_table_name">(.*)</span>')  # 匹配出榜单名称
            resu = pattern.findall(honor.text)
            print('用例成功，{0}详情:{1}'.format(data[0], resu))
            result = 'pass'
        except Exception as e:
            print('{0}接口错误!'.format(data[0]))
            log.error('ZClistinter_test:{0}'.format(e))
            result = ['fail', e]
        return result

    def input_excel_result(self, row, result):
        """
        写入结果并判断用例是否成功
        :param row:
        :param result:
        :return:
        """
        font_red = Font(color='e33e33')
        font_green = Font(color='6ca004')
        file = os.path.join(path, 'file', 'zc_interface.xlsx')
        sheetname = 'list'
        table = openpyxl.load_workbook(file)
        sheet = table[sheetname]
        if type(result) == str:
            sheet.cell(row, 6, result).font = font_green
            sheet.cell(row, 7, '')
            table.save(file)
        else:
            sheet.cell(row, 6, str(result[0])).font = font_red
            sheet.cell(row, 7, str(result[1]))
            table.save(file)
            raise NameError(result[1])


