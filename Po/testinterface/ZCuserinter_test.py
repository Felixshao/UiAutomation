import unittest, re, os
import requests
import urllib3
import ddt
import openpyxl
from openpyxl.styles import Font, colors
from bs4 import BeautifulSoup
from common.operateExcel import operateExcel
from common.log import Logger
from config.getProjectPath import get_project_path

log = Logger('Po.testinterface.ZCuserinter_test').get_logger()
s = requests.Session()  # 定义一个全局session对象
login = operateExcel(file_name='zc_interface.xlsx', sheet_name='login').get_excel_list()[0]
user = operateExcel(file_name='zc_interface.xlsx', sheet_name='user').get_excel_list()
path = get_project_path()


@ddt.ddt
class ZCuserinter_test(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        log.info('---------------------------- ZCuserinter_test start ---------------------------------------------')
        urllib3.disable_warnings()
        logins = s.post(login[1], data=eval(login[3]), verify=False)
        print(logins.text)

    @classmethod
    def tearDownClass(cls):
        log.info('---------------------------- ZCuserinter_test end ---------------------------------------------')

    @ddt.data(user[0])
    def test1_user_income(self, data):
        """用户基本资料接口"""
        data[4] = eval(data[4])
        result,  basic = self.check_hororlist_inter(data)
        self.input_excel_result(0+2, result=result)
        print(result, basic)
        soup = BeautifulSoup(basic, 'lxml')
        dict = {}
        dict['账号'] = soup.find('p', class_='bas-per-account').get_text()  # 账号
        dict['身份证认证状态'] = soup.find('span', class_='upfail').get_text()  # 身份证认证状态
        dict['账户余额'] = soup.find('span', style='color: #f99c00;').get_text()  # 账户余额
        dict['昵称'] = soup.find('input', id="nickname")['value']  # 昵称
        dict['qq'] = soup.find('input', id='qq')['value']  # qq
        select = soup.find_all('option', selected='selected')  # 性别
        for i in range(len(select)):
            if i == 0:
                dict['性别'] = select[i].get_text().replace('\n', '').replace(' ', '')
            elif i == 1:
                dict['生日'] = select[i].get_text().replace('\n', '').replace(' ', '')
            elif i == 2:
                dict['地址'] = select[i].get_text().replace('\n', '').replace(' ', '')
            elif i == 3:
                dict['职业'] = select[i].get_text().replace('\n', '').replace(' ', '')
            elif i == 4:
                dict['测试经验'] = select[i].get_text().replace('\n', '').replace(' ', '')
            elif i == 5:
                dict['管理经验'] = select[i].get_text().replace('\n', '').replace(' ', '')
        print('用户{0}为:{1}'.format(data[0], dict))
        log.info('用例成功，用户{0}为:{1}'.format(data[0], dict))

    # @ddt.data(user[1])
    # def test2_user_income(self, data):
    #     """用户测试设备接口"""
    #     data[4] = eval(data[4])
    #     result, basic = self.check_hororlist_inter(data)
    #     self.input_excel_result(1 + 2, result=result)
    #     soup = BeautifulSoup(basic, 'lxml')
    #     dicts = {}
    #     num = soup.find_all('span', class_='fr delete-equts')  # 查看设备个数
    #     cont = soup.find_all('span', class_='equip-cont')  # 账号
    #     it = 0
    #     for i in range(len(num)):
    #         dict = {}
    #         dict['品牌型号'] = cont[it].get_text()
    #         dict['操作系统'] = cont[it + 1].get_text()
    #         dict['运营商网络'] = cont[it + 2].get_text()
    #         dicts['设备' + str(i + 1)] = dict
    #         it += 3
    #     print('用例成功，用户设备{0}为:{1}'.format(data[0], dicts))
    #     log.info('用例成功，用户设备{0}为:{1}'.format(data[0], dicts))
    #
    # @ddt.data(user[2])
    # def test3_user_information(self, data):
    #     """用户身份信息接口"""
    #     data[4] = eval(data[4])
    #     result, basic = self.check_hororlist_inter(data)
    #     self.input_excel_result(2 + 2, result=result)
    #     print('用例成功!')
    #
    # @ddt.data(user[3])
    # def test4_financial_income(self, data):
    #     """财务收入明细接口"""
    #     data[4] = eval(data[4])
    #     result, basic = self.check_hororlist_inter(data)
    #     self.input_excel_result(3 + 2, result=result)
    #     print('用例成功!')
    #
    # @ddt.data(user[4])
    # def test5_gold_income(self, data):
    #     """金币收入明细接口"""
    #     data[4] = eval(data[4])
    #     result, basic = self.check_hororlist_inter(data)
    #     self.input_excel_result(4 + 2, result=result)
    #     print('用例成功!')
    #
    # @ddt.data(user[5])
    # def test6_invite_friends(self, data):
    #     """邀请好友接口"""
    #     data[4] = eval(data[4])
    #     result, basic = self.check_hororlist_inter(data)
    #     self.input_excel_result(5 + 2, result=result)
    #     print('用例成功!')

    def check_hororlist_inter(self, data):
        """断言榜单接口"""
        try:
            honor = s.get(url=data[1], params=data[3], verify=False)
            print(honor.status_code)
            self.assertEqual(honor.status_code, data[4]['status_code'],
                             msg='status_code:{0}'.format(honor.status_code))
            pattern = re.compile(pattern='<div class="dis_inlblo"><span class="rank_table_name">(.*)</span>')  # 匹配出榜单名称
            resu = pattern.findall(honor.text)
            print(resu)
            result = 'pass'
            return result, resu
        except Exception as e:
            print(2)
            log.error('ZClistinter_test:{0}'.format(e))
            result = ['fail', e]
        return result

    def input_excel_result(self, row, col=6, result='pass'):
        """
        写入结果并判断用例是否成功
        :param row:
        :param result:
        :return:
        """
        font_red = Font(color='e33e33')
        font_green = Font(color='6ca004')
        file = os.path.join(path, 'file', 'zc_interface.xlsx')
        sheetname = 'user'
        table = openpyxl.load_workbook(file)
        sheet = table[sheetname]
        if type(result) == str:
            sheet.cell(row, col, result).font = font_green
            sheet.cell(row, col+1, '')
            table.save(file)
        else:
            sheet.cell(row, col, str(result[0])).font = font_red
            sheet.cell(row, col+1, str(result[1]))
            table.save(file)
            print('{0}接口错误!'.format(result[1]))
            raise NameError(result[1])
